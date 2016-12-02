"""
@author: Weijian
"""
import logging as log
import time

from openpyxl import Workbook
from openpyxl.styles import Font

import worker.export.excel_fields as field

log.basicConfig(format='%(message)s',
                datefmt='%m/%d/%Y %I:%M:%S %p',
                filename='export.log',
                filemode="a",
                level=log.DEBUG)


def create_xls( search_id, namelist, search_items, individual_info ):

    wb = Workbook()

    # exact matches
    sheets = dict()
    sheets['exact'] = wb.active
    wb.active.title = "Exact"
    sheets['multiple'] = wb.create_sheet(title="Multiple")
    sheets['no matches'] = wb.create_sheet(title="No Matches")

    for worksheet in sheets.values():
        set_headers(worksheet)

    for search in (sorted(search_items, key=lambda x: x['search_name'])):

        sheet = 'no matches'

        if search['status'] == 'EXACT_MATCH':
            sheet = 'exact'
        elif search['status'] == 'MULTIPLE_MATCH':
            sheet = 'multiple'

        add_to_sheet( sheet, sheets[sheet], search, individual_info )

    filename = '%s_%s%s.xlsx' % (namelist, search_id, time.strftime('%y%m%d%H%M%S'))
    wb.save(filename)
    return filename

BASE_FONT = Font(size=8)
HEADER_FONT= Font(size=8, bold=True)

rows_processed = {'exact':1, 'multiple': 1 , 'no matches': 1 }


def add_to_sheet(sheet, ws, item, individual_info):

    rownum = rows_processed[sheet]
    row = {}

    row_height = 12

    row[field.NAME.colnum] = item['search_name']
    if 'firm' in item['search_info']:
        row[field.FIRM.colnum] = item['search_info']['firm']
    if 'reference' in item['search_info']:
        row[field.REFERENCE_ID.colnum] = item['search_info']['reference']

    if item['status'] == 'NO_MATCH':
        ws.append(row)
        rownum = rownum + 1
        _row = ws.row_dimensions[rownum]
        _row.ht = row_height
        _row.font = BASE_FONT
    else:

        for match in item['matches']:

            match_row = dict(row)

            info = individual_info.get(match['individual_id'])

            match_row[field.MATCH_NAME.colnum] = match['name']
            match_row[field.COUNTRY.colnum] = info['country']
            match_row[field.REGISTER.colnum] = info['register']

            match_row[field.MATCH_SCORE.colnum] = match['match_score']

            if info['status'] == 'A':
                match_row[field.INDIVIDUAL_STATUS.colnum] = 'Active'
            else:
                match_row[field.INDIVIDUAL_STATUS.colnum] = 'Inactive'

            statuses = get_status(info)

            row_height = max(12, 12 * len(statuses))
            match_row[field.INDIVIDUAL_REGISTRATION_STATUS.colnum] = '\n'.join( statuses )

            match_row[field.MATCH_REFERENCE_ID.colnum] = get_reference_id( info )

            link = get_link(info)
            if link:
                match_row[field.LINK.colnum] = '=HYPERLINK("%s", "%s link")' % (link, info['register'] )

            match_firms = match.get('firms', None )
            if len(match_firms) > 0:
                for mf in match_firms:

                    f = mf['firm']

                    row_height = max(12, 12 * len(statuses))
                    firm_row = dict(match_row)
                    firm_row[field.MATCH_FIRM.colnum] = f['n']

                    if mf['current']:
                        firm_row[field.FIRM_EMPLOYMENT.colnum] = 'CURRENT'
                    else:
                        firm_row[field.FIRM_EMPLOYMENT.colnum] = 'HISTORICAL'

                    firm_detail = f['d']
                    if firm_detail:
                        firm_statuses = [u"\u2022 %s" % s for s in firm_detail['status'] ]
                        row_height = max(row_height, 12 * len(firm_statuses))
                        firm_row[field.FIRM_STATUS.colnum] = '\n'.join(firm_statuses)

                    ws.append( firm_row )
                    rownum += 1
                    _row = ws.row_dimensions[rownum]
                    _row.ht = row_height
                    _row.font = BASE_FONT
            else:
                ws.append(match_row)
                rownum += 1
                _row = ws.row_dimensions[rownum]
                _row.ht = row_height
                _row.font = BASE_FONT
    #
    # save current rownum
    #
    rows_processed[sheet] = rownum


def get_reference_id(info):
    if info['register'] == 'CSA':
        return ''
    else:
        return info.get('ext_reference_id','')


def get_status(info):

    statuses = []

    if info['details']:

        for non_firm_status in info['details'].get('status',[]):
            statuses.append( u"\u2022 %s" % non_firm_status )

        if info['register'] in ['CSA','NFA']:

            for firm in info['details'].get('firms',[]):

                statuses.append( "%s" % firm['name'])
                for firm_status in firm['status']:
                    statuses.append( "\t%s"%firm_status )
    return statuses


def get_link(info):
    register_code = info['register']
    reference = info['ext_reference_id']
    return "%s/%s" % (register_code, reference)


def cell(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.font = Font(size=8, vertAlign='Top')
    return c


def set_headers(ws):
    headers = []
    for column in field.ALL_FIELDS:
        headers.append ( column.name )
        ws.column_dimensions[column.letter].width = column.width
    ws.append( headers )

    _row = ws.row_dimensions[ws.max_row]
    _row.font = HEADER_FONT
    return







